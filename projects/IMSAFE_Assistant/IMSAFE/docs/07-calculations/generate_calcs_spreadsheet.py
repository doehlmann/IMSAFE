#!/usr/bin/env python3
"""
IMSAFE Schematic Calculations Spreadsheet Generator

Generates a multi-tab Excel workbook with component sizing calculations
for each schematic section.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_header_style():
    """Return styling for header cells"""
    return {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_cell_style():
    """Return styling for data cells"""
    return {
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'alignment': Alignment(horizontal='left', vertical='center')
    }

def apply_header_style(cell):
    style = create_header_style()
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']

def apply_cell_style(cell):
    style = create_cell_style()
    cell.border = style['border']
    cell.alignment = style['alignment']

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_section_header(ws, row, title, cols=6):
    """Add a section header spanning multiple columns"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    return row + 1

def create_summary_sheet(wb):
    """Create summary/overview sheet"""
    ws = wb.active
    ws.title = "Summary"

    set_column_widths(ws, [20, 40, 15, 15, 30])

    # Title
    ws['A1'] = "IMSAFE Schematic Calculations"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = "Project: IMSAFE Flying Assistant"

    row = 5
    row = add_section_header(ws, row, "Document Overview", 5)

    headers = ['Sheet', 'Description', 'Status', 'Components', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    sheets = [
        ('01_MCU_Core', 'Crystal & decoupling calculations', 'Complete', '28', 'Per ST AN4661'),
        ('02_USB_C', 'CC resistor & ESD calculations', 'Complete', '6', 'USB 2.0 UFP'),
        ('03_Charger', 'BQ25792 sizing calculations', 'Complete', '17', 'Buck-boost power path'),
        ('04_BMS', 'Cell protection & balancing', 'Complete', '12', 'HY2120 + HY2213'),
        ('05_Regulators', '3.3V/5V regulator calculations', 'Pending', '~14', 'TPS62133/TPS62130'),
        ('Power_Budget', 'System power analysis', 'In Progress', '-', 'All sections'),
    ]

    for sheet in sheets:
        for col, value in enumerate(sheet, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def create_mcu_sheet(wb):
    """Create MCU Core calculations sheet"""
    ws = wb.create_sheet("01_MCU_Core")
    set_column_widths(ws, [25, 15, 12, 12, 15, 40])

    row = 1
    row = add_section_header(ws, row, "Section 01: MCU Core - STM32H747AGI6", 6)
    row += 1

    # HSE Crystal Calculations
    row = add_section_header(ws, row, "HSE Crystal Load Capacitor Calculation (8 MHz)", 6)

    headers = ['Parameter', 'Symbol', 'Value', 'Unit', 'Formula', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    hse_data = [
        ('Crystal Load Capacitance', 'CL', 10, 'pF', 'From datasheet', 'Typical 8MHz 3225 crystal'),
        ('PCB Stray Capacitance', 'Cstray', 3, 'pF', 'Estimate', 'Typical range 2-5 pF'),
        ('Calculated Load Cap', 'Cload', '=2*(C4-C5)', 'pF', 'CL = (C1*C2)/(C1+C2) + Cstray', 'Assuming C1=C2'),
        ('Selected Load Cap', 'C123/C124', 15, 'pF', 'Standard value', 'Round up to standard'),
        ('Effective CL', 'CL_eff', '=C7/2+C5', 'pF', 'Cload/2 + Cstray', 'Should be ≈ CL'),
        ('Feedback Resistor', 'R101', 1, 'MΩ', 'Recommended', 'Optional - limits drive'),
    ]

    for data in hse_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # LSE Crystal Calculations
    row = add_section_header(ws, row, "LSE Crystal Load Capacitor Calculation (32.768 kHz)", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    lse_data = [
        ('Crystal Load Capacitance', 'CL', 6, 'pF', 'From datasheet', 'Typical 32.768kHz crystal'),
        ('PCB Stray Capacitance', 'Cstray', 2, 'pF', 'Estimate', 'Lower for small crystal'),
        ('Calculated Load Cap', 'Cload', '=2*(C14-C15)', 'pF', 'CL = (C1*C2)/(C1+C2) + Cstray', 'Assuming C1=C2'),
        ('Selected Load Cap', 'C125/C126', 6.8, 'pF', 'Standard value', 'Closest standard'),
        ('Effective CL', 'CL_eff', '=C17/2+C15', 'pF', 'Cload/2 + Cstray', 'Should be ≈ CL'),
        ('Feedback Resistor', 'R102', 10, 'MΩ', 'Recommended', 'Higher for low-power LSE'),
    ]

    for data in lse_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Decoupling Summary
    row = add_section_header(ws, row, "Power Decoupling Summary (per ST AN4661)", 6)

    headers2 = ['Power Domain', 'Qty', 'Value', 'Package', 'Refs', 'Notes']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    decoupling_data = [
        ('VDD (digital)', 1, '4.7µF', '0805', 'C101', 'Bulk capacitor'),
        ('VDD (digital)', 10, '100nF', '0603', 'C102-C111', 'Per VDD pin'),
        ('VDDA (analog)', 4, '100nF', '0603', 'C112-C115', 'Per VDDA pin'),
        ('VDDA (analog)', 1, '1µF', '0603', 'C116', 'Analog bulk'),
        ('VDDUSB', 2, '100nF', '0603', 'C117-C118', 'USB PHY'),
        ('VCAP (LDO out)', 2, '2.2µF', '0603', 'C119-C120', 'Internal regulator'),
        ('VBAT', 1, '100nF', '0603', 'C121', 'RTC backup'),
        ('VREF+', 1, '10nF', '0603', 'C122', 'ADC reference'),
        ('TOTAL', 22, '-', '-', '-', 'Decoupling caps'),
    ]

    for data in decoupling_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def create_usbc_sheet(wb):
    """Create USB-C calculations sheet"""
    ws = wb.create_sheet("02_USB_C")
    set_column_widths(ws, [25, 15, 12, 12, 15, 40])

    row = 1
    row = add_section_header(ws, row, "Section 02: USB-C Input", 6)
    row += 1

    # CC Resistor Calculations
    row = add_section_header(ws, row, "CC Pull-Down Resistor Calculation (UFP/Sink Role)", 6)

    headers = ['Parameter', 'Symbol', 'Value', 'Unit', 'Spec', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    cc_data = [
        ('CC Pull-Down (UFP)', 'Rd', 5.1, 'kΩ', 'USB-C Spec', 'Identifies device as sink'),
        ('Tolerance', '-', '±10%', '%', 'USB-C Spec', '4.59k - 5.61k acceptable'),
        ('Selected Value', 'R201/R202', 5.1, 'kΩ', '-', 'Standard 1% resistor'),
        ('Voltage at CC (3A)', 'Vcc', '~1.6', 'V', 'Rp=10k/3A', 'DFP uses 10k for 3A'),
        ('Voltage at CC (1.5A)', 'Vcc', '~0.9', 'V', 'Rp=22k/1.5A', 'DFP uses 22k for 1.5A'),
    ]

    for data in cc_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # ESD Protection
    row = add_section_header(ws, row, "ESD Protection (USBLC6-2SC6)", 6)

    headers2 = ['Parameter', 'Symbol', 'Value', 'Unit', 'Requirement', 'Notes']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    esd_data = [
        ('ESD Rating (Contact)', '-', '±8', 'kV', 'IEC 61000-4-2 L4', 'Exceeds requirement'),
        ('ESD Rating (Air)', '-', '±15', 'kV', 'IEC 61000-4-2 L4', 'Exceeds requirement'),
        ('Line Capacitance', 'Cline', 3.5, 'pF', '<10pF for USB 2.0', 'USB 2.0 compatible'),
        ('Clamping Voltage', 'Vclamp', '~7', 'V', '@8kV', 'Protects 3.3V logic'),
        ('Leakage Current', 'Ileak', '<1', 'µA', '-', 'Negligible'),
    ]

    for data in esd_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # VBUS Filtering
    row = add_section_header(ws, row, "VBUS Input Filtering", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    vbus_data = [
        ('VBUS Voltage', 'VBUS', 5, 'V', 'USB Spec', 'Nominal'),
        ('VBUS Tolerance', '-', '±5%', '%', 'USB Spec', '4.75V - 5.25V'),
        ('HF Filter Cap', 'C201', 100, 'nF', '-', 'High-frequency noise'),
        ('Bulk Cap', 'C202', 10, 'µF', '-', 'Inrush/transient'),
        ('Bulk Cap Voltage', '-', 16, 'V', '>5V', '3x margin'),
    ]

    for data in vbus_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def create_charger_sheet(wb):
    """Create Charger calculations sheet"""
    ws = wb.create_sheet("03_Charger")
    set_column_widths(ws, [28, 15, 12, 12, 15, 40])

    row = 1
    row = add_section_header(ws, row, "Section 03: Battery Charger - BQ25792", 6)
    row += 1

    # System Parameters
    row = add_section_header(ws, row, "System Parameters", 6)

    headers = ['Parameter', 'Symbol', 'Value', 'Unit', 'Formula', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    sys_data = [
        ('Battery Configuration', '-', '2S', '-', '-', '2 cells in series'),
        ('Cell Voltage (nom)', 'Vcell', 3.7, 'V', '-', 'Li-ion nominal'),
        ('Cell Voltage (max)', 'Vcell_max', 4.2, 'V', '-', 'Full charge'),
        ('Cell Voltage (min)', 'Vcell_min', 3.0, 'V', '-', 'Cutoff'),
        ('Pack Voltage (nom)', 'Vbat', 7.4, 'V', '=2*C6', '2S nominal'),
        ('Pack Voltage (max)', 'Vbat_max', 8.4, 'V', '=2*C7', '2S full'),
        ('Pack Capacity', 'Cbat', 16000, 'mAh', '-', 'Per spec'),
        ('USB Input Voltage', 'Vbus', 5, 'V', '-', 'USB standard'),
    ]

    for data in sys_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Charge Time Calculation
    row = add_section_header(ws, row, "Charge Time Estimates", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    charge_data = [
        ('USB Power (5V/3A)', 'Pusb', 15, 'W', '=5*3', 'Typical USB-C'),
        ('Charger Efficiency', 'η', 90, '%', '-', 'Buck-boost typical'),
        ('Available Power', 'Pcharge', 13.5, 'W', '=C18*C19/100', 'At battery'),
        ('Charge Current @2A', 'Ichg_2A', 2, 'A', '-', 'Conservative'),
        ('Charge Current @3A', 'Ichg_3A', 3, 'A', '-', 'Moderate'),
        ('Charge Current @5A', 'Ichg_5A', 5, 'A', '-', 'Maximum'),
        ('Time @2A', 't_2A', 8.0, 'hrs', '=C12/C22/1000', '16000mAh/2000mA'),
        ('Time @3A', 't_3A', 5.3, 'hrs', '=C12/C23/1000', '16000mAh/3000mA'),
        ('Time @5A', 't_5A', 3.2, 'hrs', '=C12/C24/1000', '16000mAh/5000mA'),
    ]

    for data in charge_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Current Sense Resistors
    row = add_section_header(ws, row, "Current Sense Resistor Calculations", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    sense_data = [
        ('Input Current (max)', 'Iin_max', 3.5, 'A', '-', 'USB + margin'),
        ('Input Sense Resistor', 'R301', 10, 'mΩ', 'BQ25792 typical', 'AC sense'),
        ('Input Sense Voltage', 'Vsense_in', 35, 'mV', '=C33*C34', 'At max current'),
        ('Input Sense Power', 'Psense_in', 122.5, 'mW', '=C33^2*C34/1000', 'Power loss'),
        ('Battery Current (max)', 'Ibat_max', 5, 'A', '-', 'Max charge'),
        ('Battery Sense Resistor', 'R302', 5, 'mΩ', 'BQ25792 typical', 'Lower for efficiency'),
        ('Battery Sense Voltage', 'Vsense_bat', 25, 'mV', '=C38*C39', 'At max current'),
        ('Battery Sense Power', 'Psense_bat', 125, 'mW', '=C38^2*C39/1000', 'Power loss'),
    ]

    for data in sense_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Inductor Selection
    row = add_section_header(ws, row, "Power Inductor Selection", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    inductor_data = [
        ('Switching Frequency', 'Fsw', 750, 'kHz', 'BQ25792 typical', 'Default'),
        ('Inductor Value', 'L1', 2.2, 'µH', 'BQ25792 recommended', '1-2.2µH range'),
        ('Saturation Current', 'Isat', 5, 'A', '>=Ichg_max', 'Must not saturate'),
        ('DC Resistance', 'Rdc', 15, 'mΩ', 'Typical', 'Lower is better'),
        ('Inductor Power Loss', 'PL', 375, 'mW', '=C24^2*C50', 'At 5A'),
        ('Package', '-', '5x5mm', '-', '-', 'Shielded preferred'),
    ]

    for data in inductor_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # I2C Pull-ups
    row = add_section_header(ws, row, "I2C Pull-Up Resistor Calculation", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    i2c_data = [
        ('I2C Speed', '-', 400, 'kHz', '-', 'Fast mode'),
        ('Bus Capacitance (est)', 'Cbus', 50, 'pF', '-', 'Short traces'),
        ('Rise Time (max)', 'tr', 300, 'ns', 'I2C FM spec', 'Fast mode'),
        ('Pull-Up (max)', 'Rp_max', '=0.3*C59/C60*1e9', 'kΩ', 'tr = 0.3 * Rp * Cb', '~6kΩ'),
        ('Pull-Up (min)', 'Rp_min', '=(3.3-0.4)/3', 'kΩ', 'Vol=0.4V, Iol=3mA', '~1kΩ'),
        ('Selected Pull-Up', 'R305/R306', 4.7, 'kΩ', '-', 'Standard value'),
    ]

    for data in i2c_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def create_bms_sheet(wb):
    """Create BMS calculations sheet"""
    ws = wb.create_sheet("04_BMS")
    set_column_widths(ws, [28, 15, 12, 12, 15, 40])

    row = 1
    row = add_section_header(ws, row, "Section 04: Battery Management System (BMS)", 6)
    row += 1

    # System Configuration
    row = add_section_header(ws, row, "Battery Configuration", 6)

    headers = ['Parameter', 'Symbol', 'Value', 'Unit', 'Formula', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    config_data = [
        ('Cell Configuration', '-', '2S', '-', '-', '2 cells in series'),
        ('Cell Voltage (nom)', 'Vcell', 3.7, 'V', '-', 'Li-ion nominal'),
        ('Cell Voltage (max)', 'Vcell_max', 4.2, 'V', '-', 'Full charge'),
        ('Cell Voltage (min)', 'Vcell_min', 3.0, 'V', '-', 'Discharge cutoff'),
        ('Pack Voltage (nom)', 'Vpack', 7.4, 'V', '=2*3.7', '2S nominal'),
        ('Pack Voltage (max)', 'Vpack_max', 8.4, 'V', '=2*4.2', '2S full charge'),
        ('Pack Capacity', 'Cbat', 16000, 'mAh', '-', 'Per spec'),
    ]

    for data in config_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Protection Thresholds
    row = add_section_header(ws, row, "HY2120 Protection Thresholds", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    prot_data = [
        ('Overcharge Threshold', 'Voc', 4.25, 'V', '-', 'Per cell, CB version'),
        ('Overcharge Release', 'Vocr', 4.15, 'V', '-', 'Hysteresis'),
        ('Over-discharge Threshold', 'Vod', 2.8, 'V', '-', 'Per cell'),
        ('Over-discharge Release', 'Vodr', 3.0, 'V', '-', 'Hysteresis'),
        ('Overcurrent (discharge)', 'Ioc_dis', 200, 'mV', '-', 'Across sense resistor'),
        ('Overcurrent (charge)', 'Ioc_chg', 100, 'mV', '-', 'Lower for charging'),
        ('Short Circuit', 'Isc', 400, 'mV', '-', 'Fast response'),
    ]

    for data in prot_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Current Sense Calculation
    row = add_section_header(ws, row, "Current Sense Resistor Calculation", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    sense_data = [
        ('Overcurrent Threshold', 'Voc', 200, 'mV', '-', 'HY2120 typical'),
        ('Desired Trip Current', 'Itrip', 40, 'A', '-', 'Max continuous'),
        ('Sense Resistor', 'R403', '=200/40', 'mΩ', 'Rsense = Voc/Itrip', '5mΩ calculated'),
        ('Selected Value', 'R403', 5, 'mΩ', '-', 'Standard 1% value'),
        ('Actual Trip Current', 'Itrip_act', '=200/5', 'A', 'I = Voc/R', '40A'),
        ('Power @ Trip', 'Psense', '=40^2*5/1000', 'W', 'P = I²R', '8W peak'),
        ('Power @ 5A Normal', 'Psense_5A', '=5^2*5/1000', 'mW', 'P = I²R', '125mW'),
        ('Resistor Package', '-', '1206', '-', '-', '0.5W rating'),
    ]

    for data in sense_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # Cell Balancing Calculation
    row = add_section_header(ws, row, "Cell Balancing (HY2213-BB3A)", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    balance_data = [
        ('Balance Start Voltage', 'Vbal', 4.2, 'V', '-', 'HY2213-BB3A threshold'),
        ('Balance Voltage Hyst', 'Vbal_h', 50, 'mV', '-', 'Hysteresis'),
        ('Desired Bleed Current', 'Ibal', 68, 'mA', '-', 'Moderate balance rate'),
        ('Balance Resistor', 'Rbal', '=4.2/0.068', 'Ω', 'R = V/I', '~62Ω'),
        ('Selected Value', 'R404/R405', 62, 'Ω', '-', 'Standard E24 value'),
        ('Actual Bleed Current', 'Ibal_act', '=4.2/62*1000', 'mA', 'I = V/R', '67.7mA'),
        ('Power Dissipation', 'Pbal', '=4.2^2/62*1000', 'mW', 'P = V²/R', '284mW'),
        ('Resistor Package', '-', '0805', '-', '-', '0.25W rating OK'),
        ('Max Balance Time', 't_bal', '=500/67.7', 'hrs', 'ΔQ / Ibal', '~7.4h for 500mAh diff'),
    ]

    for data in balance_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # MOSFET Selection
    row = add_section_header(ws, row, "MOSFET Selection (AO3400A)", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    mosfet_data = [
        ('Drain-Source Voltage', 'Vds', 30, 'V', '-', 'Exceeds 8.4V pack'),
        ('Continuous Current', 'Id', 5.7, 'A', '-', 'At 25°C'),
        ('On Resistance', 'Rds_on', 28, 'mΩ', '-', 'At Vgs=4.5V'),
        ('Gate Threshold', 'Vgs_th', 1.2, 'V', '-', 'Typical'),
        ('HY2120 Gate Drive', 'Vgate', '~3', 'V', '-', 'From OD/OC pins'),
        ('FETs in Series', 'n', 2, '-', '-', 'Back-to-back per function'),
        ('Total Rds (per path)', 'Rds_total', '=28*2', 'mΩ', 'n * Rds_on', '56mΩ'),
        ('Conduction Loss @5A', 'Pcond', '=5^2*56/1000', 'W', 'P = I² * R', '1.4W'),
        ('Gate Resistor', 'R401/R402', 1, 'kΩ', '-', 'Limits gate current'),
    ]

    for data in mosfet_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # BOM Summary
    row = add_section_header(ws, row, "Section 04 BOM Summary", 6)

    headers2 = ['Reference', 'Value', 'Package', 'Qty', 'JLCPCB', 'Description']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    bom_data = [
        ('U4', 'HY2120-CB', 'SOT-23-6', 1, 'C113632', '2S protection IC'),
        ('U5, U6', 'HY2213-BB3A', 'SOT-23-6', 2, 'C113633', 'Cell balancer'),
        ('Q1-Q4', 'AO3400A', 'SOT-23', 4, 'C20917', 'N-ch MOSFET'),
        ('C401-C403', '100nF', '0603', 3, 'C14663', 'Bypass caps'),
        ('R401, R402', '1kΩ', '0603', 2, 'C21190', 'Gate resistors'),
        ('R403', '5mΩ', '1206', 1, 'C2933641', 'Current sense'),
        ('R404, R405', '62Ω', '0805', 2, 'C17828', 'Balance resistors'),
        ('TOTAL', '-', '-', 15, '-', '4 ICs + 11 passives'),
    ]

    for data in bom_data:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def create_power_budget_sheet(wb):
    """Create Power Budget calculations sheet"""
    ws = wb.create_sheet("Power_Budget")
    set_column_widths(ws, [30, 12, 12, 12, 12, 35])

    row = 1
    row = add_section_header(ws, row, "System Power Budget Analysis", 6)
    row += 1

    headers = ['Subsystem', 'Typ (mA)', 'Max (mA)', 'Voltage', 'Power (mW)', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    # 3.3V Rail
    power_3v3 = [
        ('STM32H747 (M7 active)', 150, 300, 3.3, '=B4*D4', 'M7 @ 480MHz'),
        ('STM32H747 (M4 active)', 50, 100, 3.3, '=B5*D5', 'M4 @ 240MHz'),
        ('BMP390 (barometric)', 0.7, 1, 3.3, '=B6*D6', 'Continuous'),
        ('ICM-42688-P (IMU)', 1, 3, 3.3, '=B7*D7', 'Accel+gyro active'),
        ('MMC5983MA (mag)', 0.5, 1, 3.3, '=B8*D8', 'Continuous'),
        ('Si7021 (temp/humid)', 0.15, 0.2, 3.3, '=B9*D9', 'Low power'),
        ('MCP23017 (I/O exp)', 1, 1.5, 3.3, '=B10*D10', 'Active'),
        ('ESP32-C3 (BLE)', 20, 350, 3.3, '=B11*D11', 'TX peak'),
        ('microSD (active)', 30, 100, 3.3, '=B12*D12', 'Write operation'),
        ('3.3V Subtotal', '=SUM(B4:B12)', '=SUM(C4:C12)', 3.3, '=SUM(E4:E12)', '-'),
    ]

    for data in power_3v3:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # 5V Rail
    headers2 = ['Subsystem', 'Typ (mA)', 'Max (mA)', 'Voltage', 'Power (mW)', 'Notes']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    power_5v = [
        ('WS2812B-2020 (16 LEDs)', 50, 192, 5, '=B17*D17', '12mA max each'),
        ('MICS-4514 (heater)', 80, 100, 5, '=B18*D18', 'Heater active'),
        ('Display backlight', 50, 130, 5, '=B19*D19', 'PWM controlled'),
        ('Buzzer (active)', 0, 30, 5, '=B20*D20', 'Intermittent'),
        ('5V Subtotal', '=SUM(B17:B20)', '=SUM(C17:C20)', 5, '=SUM(E17:E20)', '-'),
    ]

    for data in power_5v:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    row += 1

    # System Totals
    row = add_section_header(ws, row, "System Totals", 6)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    totals = [
        ('3.3V Rail (from VSYS)', '=B13', '=C13', 3.3, '=E13', '-'),
        ('5V Rail (from VSYS)', '=B21', '=C21', 5, '=E21', '-'),
        ('Regulator Losses (~10%)', '-', '-', '-', '=(E26+E27)*0.1', 'Efficiency loss'),
        ('TOTAL from VSYS', '-', '-', '~7.4', '=E26+E27+E28', 'Typ/Max'),
        ('Battery Current (typ)', '-', '-', 7.4, '=E29/7400*1000', 'mA from battery'),
        ('Battery Life (16Ah)', '-', '-', '-', '=16000/E30', 'Hours'),
    ]

    for data in totals:
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_cell_style(cell)
        row += 1

    return ws

def main():
    """Main function to generate the workbook"""
    wb = Workbook()

    # Create all sheets
    create_summary_sheet(wb)
    create_mcu_sheet(wb)
    create_usbc_sheet(wb)
    create_charger_sheet(wb)
    create_bms_sheet(wb)
    create_power_budget_sheet(wb)

    # Save workbook
    output_path = "IMSAFE_Schematic_Calculations.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
